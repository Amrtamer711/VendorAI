from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import os
from slack_sdk import WebClient
from config import SLACK_BOT_TOKEN
import requests
import re
import subprocess
from clients import web_client

def download_slack_file(file_info, user, user_file_store):
    file_name = file_info["name"]
    file_url = file_info["url_private"]
    headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}
    safe_file_name = re.sub(r'[^\w\-_\.]', '_', file_name)
    local_path = f"./uploads/{safe_file_name}"
    os.makedirs("./uploads", exist_ok=True)

    response = requests.get(file_url, headers=headers)
    with open(local_path, "wb") as f_out:
        f_out.write(response.content)

    # Store the file path
    if "vendor" in file_name.lower():
        user_file_store[user]["vendor"] = local_path
        print(f"📥 Vendor file saved for {user}: {local_path}")
    elif "soa" in file_name.lower():
        user_file_store[user]["soa"] = local_path
        print(f"📥 SOA file saved for {user}: {local_path}")
    else:
        print(f"⚠️ Unknown file type: {file_name}")

    return local_path

def send_file_to_user(channel, file_path, say, title="Reconciliation Report"):
    try:
        with open(file_path, "rb") as f:
            response = web_client.files_upload_v2(
                channel=channel,
                initial_comment="✅ Here is your reconciled Excel file:",
                file_uploads=[{
                    "filename": os.path.basename(file_path),
                    "title": title,
                    "file": f
                }]
            )
        print(f"✅ File sent to Slack")
    except Exception as e:
        say("⚠️ Failed to send the Excel file back.")
        print(f"❌ File upload failed: {e}")


def cleanup_files(*paths):
    for path in paths:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            print(f"⚠️ Cleanup failed: {e}")

def inject_recon_values_to_excel(df_fully, df_partial, df_unmatched, unbooked_difference, vendor_claimed_total, output_path="filled_recon.xlsx"):
    template_path = "./templates/recon_template.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    # Totals
    total_booked = round(df_fully["SOA Remaining"].sum(), 2)
    total_partial = round(sum(x["Difference"] for x in unbooked_difference), 2)
    total_unmatched = round(df_unmatched["Amount"].sum(), 2)
    adjusted_books_balance = round(total_booked + total_unmatched + total_partial, 2)
    difference = round(vendor_claimed_total - adjusted_books_balance, 2)

    # Inject MMG core totals
    ws["D11"] = total_booked
    ws["D14"] = total_partial + total_unmatched

    # Write unmatched
    start_row = 15
    ws[f"A{start_row}"] = "Invoice not booked"
    ws[f"A{start_row}"].font = Font(bold=True)
    start_row += 1
    for item in df_unmatched.to_dict(orient="records"):
        ws[f"A{start_row}"] = item["Invoice Number"]
        ws[f"C{start_row}"] = item["Amount"]
        start_row += 1

    # Write payment not booked
    start_row += 1
    ws[f"A{start_row}"] = "Payment not booked"
    ws[f"A{start_row}"].font = Font(bold=True)
    start_row += 1
    for item in unbooked_difference:
        desc = f"Payment not recorded as per SOA ({item['Invoice Number']})"
        ws[f"A{start_row}"] = desc
        ws[f"C{start_row}"] = item["Difference"]
        start_row += 1

    # === Borders ===
    left_right = Side(style="thin", color="000000")
    bottom = Side(style="thin", color="000000")

    # Apply vertical borders to A-D from row 14 up to current row
    for row in range(14, start_row):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = Border(left=left_right, right=left_right)

    # Add horizontal bottom border after final data row (A–C), D untouched here
    for col in ["A", "B", "C"]:
        ws[f"{col}{start_row - 1}"].border = Border(left=left_right, right=left_right, bottom=bottom)
    ws[f"D{start_row - 1}"].border = Border(left=left_right, right=left_right)

    # === Totals block (spaced rows with boxed D + horizontal border on A–C) ===
    labels = [
        ("Vendor Claimed Total", vendor_claimed_total, False),
        ("Adjusted Books Balance", adjusted_books_balance, False),
        ("Difference", difference, True),
    ]

    total_start = start_row
    for i, (label, value, is_red) in enumerate(labels):
        row = total_start + (i * 2)

        ws[f"A{row}"] = label
        ws[f"D{row}"] = value

        if is_red:
            bold_font = Font(bold=True, color="FF0000")
        else:
            bold_font = Font(bold=True)
        ws[f"A{row}"].font = bold_font
        ws[f"D{row}"].font = bold_font

        ws[f"D{row}"].border = Border(top=bottom, bottom=bottom, left=left_right, right=left_right)
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = Border(bottom=bottom)

        # Only draw blank-row borders if not the last row ("Difference")
        if i < len(labels) - 1:
            next_row = row + 1
            for col in ["A", "B", "C"]:
                ws[f"{col}{next_row}"].border = Border(bottom=bottom)
        ws[f"D{next_row}"].border = Border(top=bottom, bottom=bottom, left=left_right, right=left_right)

    wb.save(output_path)
    print(f"✅ Excel populated with aligned totals and formatting: {output_path}")

def convert_excel_to_pdf_mac(input_path):
    output_dir = os.path.dirname(input_path)
    try:
        subprocess.run([
            "soffice", "--headless", "--convert-to", "pdf", "--outdir",
            output_dir, input_path
        ], check=True)
        return os.path.splitext(input_path)[0] + ".pdf"
    except subprocess.CalledProcessError as e:
        print("❌ LibreOffice conversion failed:", e)
        return None

