from flask import Flask, request, jsonify
from slack_bolt import App
from slack_bolt.adapter.flask import SlackRequestHandler
import openai
from config import SLACK_BOT_TOKEN, SLACK_SIGNING_SECRET, OPENAI_API_KEY
from openai import OpenAI
import time
import threading
from slack_sdk import WebClient
import re
from database import log_message, init_db
import requests
import os
from datetime import datetime, timedelta
from collections import defaultdict
import subprocess
import pandas as pd
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from collections import defaultdict
from slack_sdk.errors import SlackApiError

# Initialize Slack app and Flask server
slack_app = App(token=SLACK_BOT_TOKEN, signing_secret=SLACK_SIGNING_SECRET)
flask_app = Flask(__name__)
handler = SlackRequestHandler(slack_app)
client = OpenAI(api_key=OPENAI_API_KEY, timeout=3600)
web_client = WebClient(token=SLACK_BOT_TOKEN)

# === Global Stores ===
user_history = defaultdict(list)
user_file_store = defaultdict(dict)
MAX_HISTORY_LENGTH = 6

# === Utilities ===
def download_slack_file(file_info, user):
    file_name = file_info["name"]
    file_url = file_info["url_private"]
    headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}
    safe_file_name = re.sub(r'[^\w\-_\.]', '_', file_name)
    local_path = f"./uploads/{safe_file_name}"
    os.makedirs("./uploads", exist_ok=True)

    response = requests.get(file_url, headers=headers)
    with open(local_path, "wb") as f_out:
        f_out.write(response.content)

    # Store the file path
    if "vendor" in file_name.lower():
        user_file_store[user]["vendor"] = local_path
        print(f"📥 Vendor file saved for {user}: {local_path}")
    elif "soa" in file_name.lower():
        user_file_store[user]["soa"] = local_path
        print(f"📥 SOA file saved for {user}: {local_path}")
    else:
        print(f"⚠️ Unknown file type: {file_name}")

    return local_path

def send_file_to_user(channel, file_path, say, title="Reconciliation Report"):
    try:
        with open(file_path, "rb") as f:
            response = web_client.files_upload_v2(
                channel=channel,
                initial_comment="✅ Here is your reconciled Excel file:",
                file_uploads=[{
                    "filename": os.path.basename(file_path),
                    "title": title,
                    "file": f
                }]
            )
        print(f"✅ File sent to Slack")
    except Exception as e:
        say("⚠️ Failed to send the Excel file back.")
        print(f"❌ File upload failed: {e}")

def inject_recon_values_to_excel(df_fully, df_partial, df_unmatched, unbooked_difference, vendor_claimed_total, output_path="filled_recon.xlsx"):
    template_path = "./templates/recon_template.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    # Totals
    total_booked = round(df_fully["SOA Remaining"].sum(), 2)
    total_partial = round(sum(x["Difference"] for x in unbooked_difference), 2)
    total_unmatched = round(df_unmatched["Amount"].sum(), 2)
    adjusted_books_balance = round(total_booked + total_unmatched + total_partial, 2)
    difference = round(vendor_claimed_total - adjusted_books_balance, 2)

    # Inject MMG core totals
    ws["D11"] = total_booked
    ws["D14"] = total_partial + total_unmatched

    # Write unmatched
    start_row = 15
    ws[f"A{start_row}"] = "Invoice not booked"
    ws[f"A{start_row}"].font = Font(bold=True)
    start_row += 1
    for item in df_unmatched.to_dict(orient="records"):
        ws[f"A{start_row}"] = item["Invoice Number"]
        ws[f"C{start_row}"] = item["Amount"]
        start_row += 1

    # Write payment not booked
    start_row += 1
    ws[f"A{start_row}"] = "Payment not booked"
    ws[f"A{start_row}"].font = Font(bold=True)
    start_row += 1
    for item in unbooked_difference:
        desc = f"Payment not recorded as per SOA ({item['Invoice Number']})"
        ws[f"A{start_row}"] = desc
        ws[f"C{start_row}"] = item["Difference"]
        start_row += 1

    # === Borders ===
    left_right = Side(style="thin", color="000000")
    bottom = Side(style="thin", color="000000")

    # Apply vertical borders to A-D from row 14 up to current row
    for row in range(14, start_row):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = Border(left=left_right, right=left_right)

    # Add horizontal bottom border after final data row (A–C), D untouched here
    for col in ["A", "B", "C"]:
        ws[f"{col}{start_row - 1}"].border = Border(left=left_right, right=left_right, bottom=bottom)
    ws[f"D{start_row - 1}"].border = Border(left=left_right, right=left_right)

    # === Totals block (spaced rows with boxed D + horizontal border on A–C) ===
    labels = [
        ("Vendor Claimed Total", vendor_claimed_total, False),
        ("Adjusted Books Balance", adjusted_books_balance, False),
        ("Difference", difference, True),
    ]

    total_start = start_row
    for i, (label, value, is_red) in enumerate(labels):
        row = total_start + (i * 2)

        ws[f"A{row}"] = label
        ws[f"D{row}"] = value

        if is_red:
            bold_font = Font(bold=True, color="FF0000")
        else:
            bold_font = Font(bold=True)
        ws[f"A{row}"].font = bold_font
        ws[f"D{row}"].font = bold_font

        ws[f"D{row}"].border = Border(top=bottom, bottom=bottom, left=left_right, right=left_right)
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].border = Border(bottom=bottom)

        # Only draw blank-row borders if not the last row ("Difference")
        if i < len(labels) - 1:
            next_row = row + 1
            for col in ["A", "B", "C"]:
                ws[f"{col}{next_row}"].border = Border(bottom=bottom)
        ws[f"D{next_row}"].border = Border(top=bottom, bottom=bottom, left=left_right, right=left_right)

    wb.save(output_path)
    print(f"✅ Excel populated with aligned totals and formatting: {output_path}")

def prepare_vendor_df(df):
    df = df[["Posting Date", "External Document No.", "Amount (LCY)", "Remaining Amt. (LCY)"]].copy()
    df.rename(columns={
        "Posting Date": "Date",
        "External Document No.": "Invoice Number",
        "Amount (LCY)": "Amount",
        "Remaining Amt. (LCY)": "Remaining Amount"
    }, inplace=True)

    df["Date"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime('%Y-%m-%d')
    df["Invoice Number"] = df["Invoice Number"].astype(str).str.strip()
    df["Amount"] = pd.to_numeric(df["Amount"], errors='coerce').abs()
    df["Remaining Amount"] = pd.to_numeric(df["Remaining Amount"], errors='coerce').abs()
    return df.reset_index(drop=True)

def prepare_soa_df(df):
    df = df[df["Date"].astype(str).str.upper() != "TOTAL"].copy()
    df["Date"] = pd.to_datetime(df["Date"], format="%d/%b/%Y", errors='coerce').dt.strftime('%Y-%m-%d')
    df["Invoice Number"] = df["Invoice Number"].astype(str).str.strip()
    df["Amount"] = pd.to_numeric(df["Amount"], errors='coerce').abs()
    df["Remaining Amount"] = pd.to_numeric(df["Remaining Amount"], errors='coerce').abs()
    df.rename(columns={"index": "Row"}, inplace=True)
    return df.reset_index(drop=True)

def reconcile_and_report(df_vendor, df_soa, say):
    matched_fully = []
    matched_partial = []
    unmatched_rows = []

    agreed_to_pay = []
    unbooked_difference = []

    for _, soa_row in df_soa.iterrows():
        raw_invoice = str(soa_row["Invoice Number"]).strip()
        candidate = raw_invoice.split()[0]
        match = df_vendor[df_vendor["Invoice Number"] == candidate]

        if not match.empty:
            vendor_row = match.iloc[0]
            soa_remain = soa_row["Remaining Amount"]
            vendor_remain = vendor_row["Remaining Amount"]
            row_data = {
                "Row": soa_row["Row"],
                "Invoice Number": candidate,
                "Date_soa": soa_row["Date"],
                "Date_vendor": vendor_row["Date"],
                "SOA Amount": soa_row["Amount"],
                "Vendor Amount": vendor_row["Amount"],
                "SOA Remaining": soa_remain,
                "Vendor Remaining": vendor_remain
            }

            if abs(soa_remain - vendor_remain) < 1:
                matched_fully.append(row_data)
                agreed_to_pay.append(soa_remain)
            else:
                matched_partial.append(row_data)
                unbooked_difference.append({
                    "Invoice Number": candidate,
                    "Difference": round(max(soa_remain - vendor_remain, 0), 2)
                })
        else:
            unmatched_rows.append({
                "Row": soa_row["Row"],
                "Invoice Number": raw_invoice,
                "Amount": soa_row["Amount"],
                "Remaining Amount": soa_row["Remaining Amount"],
                "Date_soa": soa_row["Date"]
            })

    # Convert all to DataFrames
    df_fully = pd.DataFrame(matched_fully)
    df_partial = pd.DataFrame(matched_partial)
    df_unmatched = pd.DataFrame(unmatched_rows)

    say(f"✅ *PAYMENTS BOOKED (MATCHED):*\n```{df_fully.to_string(index=False)}```")
    say(f"Total agreed to pay: `{sum(agreed_to_pay):,.2f}`")

    if not df_partial.empty:
        say(f"🔍 *PAYMENTS NOT BOOKED (MISMATCHED AMOUNTS):*\n```{df_partial.to_string(index=False)}```")
        say(f"Payment not booked total: `{sum(x['Difference'] for x in unbooked_difference):,.2f}`")

    if not df_unmatched.empty:
        say(f"❌ *INVOICES NOT BOOKED/RECEIVED:*\n```{df_unmatched.to_string(index=False)}```")
        say(f"Unmatched amount total: `{df_unmatched['Amount'].sum():,.2f}`")

    # Optional: Return the 3 tables for downstream use (e.g., Excel injection)
    return df_fully, df_partial, df_unmatched, unbooked_difference, 1000000

def extract_claimed_total_from_pdf(file_path: str) -> float:
    system_prompt = """You are a financial assistant at MMG. Your job is to extract the total claimed by the vendor from a Statement of Account (SOA) PDF.
The total is usually labeled something like "Total", "Outstanding", "Balance Due", or similar, and appears at the bottom of the document.
Follow these rules:
1. Respond with only the number — no explanation, no labels, no formatting.
2. Use dot as the decimal separator.
3. Strip any currency symbols, commas, or spaces.
4. If there are multiple totals, select the one that reflects the final amount due from MMG to the vendor."""

    try:
        file_upload = client.files.create(
            file=open(file_path, 'rb'),
            purpose="user_data"
        )
        file_id = file_upload.id

        print(f"📤 Uploaded for claimed total extraction: {file_id}")

        response = client.responses.create(
            model="gpt-4.1",
            input=[
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "input_file", "file_id": file_id},
                        {"type": "input_text", "text": "Extract only the claimed total amount."}
                    ]
                }
            ],
            store=False  # Do not store the file permanently
        )
        client.files.delete(file_id)

        output = response.output_text.strip()
        return float(output)

    except Exception as e:
        print(f"❌ Claimed total extraction failed: {e}")
        return 0.0  # fallback or raise if preferred

def process_dirty_vendor_reconciliation(vendor_path, soa_path, say, channel, user_comments=None):
    # Convert Excel to PDF if needed
    if soa_path.endswith(".xlsx"):
        soa_path_pdf = convert_excel_to_pdf_mac(soa_path)
        if not soa_path_pdf:
            say("❌ Could not convert SOA Excel to PDF.")
            return
    else:
        soa_path_pdf = soa_path

    # Extract SOA data using GPT
    soa_output = generate_gpt_table(
        user_text="Extract all invoice numbers, dates, amounts and show total at bottom as table only.",
        file_path=soa_path_pdf,
        user_comments=user_comments
    )

    try:
        df_soa = extract_table_from_gpt_output(soa_output)
        df_soa = prepare_soa_df(df_soa)
        df_vendor = pd.read_excel(vendor_path)
        df_vendor = prepare_vendor_df(df_vendor)
    except Exception as e:
        say(f"❌ Failed to prepare data: {e}")
        return
    
    vendor_claimed_total = extract_claimed_total_from_pdf(soa_path_pdf)
    
    say(f"*FULL SOA PARSED INVOICE TABLE (PLEASE CHECK):*\n```{df_soa.to_string(index=False)}```")
    say(f"Vendor Remaining Amount in Claimed Invoices: `{df_soa['Remaining Amount'].sum():,.2f}`")
    say(f"Vendor Total Remaining Claim: `{vendor_claimed_total:,.2f}`")

    df_fully, df_partial, df_unmatched, unbooked_difference, _ = reconcile_and_report(df_vendor, df_soa, say)
    inject_recon_values_to_excel(df_fully, df_partial, df_unmatched, unbooked_difference, vendor_claimed_total)
    send_file_to_user(channel, "filled_recon.xlsx", say)

    cleanup_files(vendor_path, soa_path, soa_path_pdf)


def process_clean_vendor_reconciliation(vendor_path, soa_path, say, channel):
    try:
        df_vendor = pd.read_excel(vendor_path)
        df_soa_raw = pd.read_excel(soa_path)
    except Exception as e:
        say(f"❌ Failed to read Excel files: {e}")
        return

    column_map_prompt = f"""
You are a data assistant. Your task is to rename columns in a vendor Statement of Account (SOA) Excel file.

Your goal is to map each original column to one of the following **target names** (use these names exactly):
• Date  
• Invoice Number  
• Amount  
• Remaining Amount

Here is the list of original columns:
{list(df_soa_raw.columns)}

Follow these strict rules:
1. Only return a valid Python dictionary of column mappings. No explanation or text outside the dictionary.
2. Only include "Remaining Amount" in your mapping **if the original columns clearly contain a second distinct amount-related column**.
3. If there's only one amount column, map it to "Amount" only — do **not** include "Remaining Amount". We will handle duplication in the backend.
4. Do not guess or make assumptions about missing fields. Only map what is clearly present.
5. Use each mapped column name only once.

✅ Example (if two amount columns):
{{"Posting Date": "Date", "Ext Doc No": "Invoice Number", "Amount (LCY)": "Amount", "Remaining (LCY)": "Remaining Amount"}}

✅ Example (if only one amount column):
{{"Posting Date": "Date", "Ext Doc No": "Invoice Number", "Amount (LCY)": "Amount"}}
"""


    try:
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "You are a helpful Python assistant."},
                {"role": "user", "content": column_map_prompt}
            ]
        )
        column_map = eval(response.choices[0].message.content.strip())
    except Exception as e:
        say(f"❌ GPT column mapping failed: {e}")
        return

    try:
        df_soa = df_soa_raw.rename(columns=column_map)

        # Only keep relevant columns
        required_cols = ["Date", "Invoice Number", "Amount", "Remaining Amount"]
        available_cols = df_soa.columns.tolist()

        # If "Remaining Amount" is not present, duplicate "Amount"
        if "Remaining Amount" not in available_cols and "Amount" in available_cols:
            df_soa["Remaining Amount"] = df_soa["Amount"]
            say("ℹ️ Only one amount column was detected. 'Remaining Amount' has been copied from 'Amount'.")
        elif "Remaining Amount" not in available_cols:
            say("❌ Neither 'Amount' nor 'Remaining Amount' found. Aborting.")
            return

        # Keep only relevant columns (preserve order)
        df_soa = df_soa[[col for col in required_cols if col in df_soa.columns]]
        df_soa = df_soa.reset_index().rename(columns={"index": "Row"})

        last_row = df_soa.iloc[-1]
        vendor_claimed_total = last_row["Remaining Amount"]
        df_soa = df_soa.iloc[:-1]  # Drop last row

        df_soa = prepare_soa_df(df_soa)
        df_vendor = prepare_vendor_df(df_vendor)

        # Show initial state of clean SOA
        say(f"*FULL SOA PARSED INVOICE TABLE (PLEASE CHECK):*\n```{df_soa.to_string(index=False)}```")
        say(f"Vendor Remaining Amount in Claimed Invoices: `{df_soa['Remaining Amount'].sum():,.2f}`")
        say(f"🧾 Using last row value as vendor claimed total: `{vendor_claimed_total:,.2f}`")

    except Exception as e:
        say(f"❌ Data preparation failed: {e}")
        return

    df_fully, df_partial, df_unmatched, unbooked_difference, vendor_claimed_total = reconcile_and_report(df_vendor, df_soa, say)
    inject_recon_values_to_excel(df_fully, df_partial, df_unmatched, unbooked_difference, vendor_claimed_total)
    send_file_to_user(channel, "filled_recon.xlsx", say)

    cleanup_files(vendor_path, soa_path)

def cleanup_files(*paths):
    for path in paths:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            print(f"⚠️ Cleanup failed: {e}")

def extract_table_from_gpt_output(output_text):
    lines = output_text.strip().splitlines()
    table_lines = [line for line in lines if line.strip().startswith("|") and line.strip().endswith("|")]
    cleaned = "\n".join(["|".join([col.strip() for col in row.strip().strip("|").split("|")]) for row in table_lines if "---" not in row])
    df = pd.read_csv(StringIO(cleaned), sep="|")
    df.columns = [col.strip() for col in df.columns]
    df.reset_index(inplace=True)
    return df

def convert_excel_to_pdf_mac(input_path):
    output_dir = os.path.dirname(input_path)
    try:
        subprocess.run([
            "soffice", "--headless", "--convert-to", "pdf", "--outdir",
            output_dir, input_path
        ], check=True)
        return os.path.splitext(input_path)[0] + ".pdf"
    except subprocess.CalledProcessError as e:
        print("❌ LibreOffice conversion failed:", e)
        return None

def markdown_to_slack(md: str) -> str:
    # Convert bold (**text** or *text*)
    md = re.sub(r'\*\*(.+?)\*\*', r'*\1*', md)
    md = re.sub(r'\*(.+?)\*', r'*\1*', md)

    # Convert italic (__text__ or _text_)
    md = re.sub(r'__(.+?)__', r'_\1_', md)
    md = re.sub(r'_(.+?)_', r'_\1_', md)

    # Convert strikethrough
    md = re.sub(r'~~(.+?)~~', r'~\1~', md)

    # Convert inline code
    md = re.sub(r'`([^`]+)`', r'`\1`', md)

    # Convert code blocks
    md = re.sub(r'```(?:\w+)?\n(.*?)\n```', r'```\1```', md, flags=re.DOTALL)

    # Convert links [text](url) -> <url|text>
    md = re.sub(r'\[(.*?)\]\((.*?)\)', r'<\2|\1>', md)

    # Convert - or * bullet points to •
    md = re.sub(r'^[\-\*]\s+', r'• ', md, flags=re.MULTILINE)

    # Convert numbered lists to dot bullets
    md = re.sub(r'^\d+\.\s+', r'• ', md, flags=re.MULTILINE)

    # Remove headings (#, ##, ###) – Slack doesn't support them
    md = re.sub(r'^#{1,6}\s+', '', md, flags=re.MULTILINE)

    # Remove extra trailing spaces
    md = re.sub(r'[ \t]+$', '', md, flags=re.MULTILINE)

    return md.strip()

def get_user_profile(user_id):
    try:
        response = web_client.users_info(user=user_id)
        profile = response["user"]["profile"]
        return {
            "display_name": profile.get("display_name"),
            "real_name": profile.get("real_name"),
            "email": profile.get("email"),
            "title": profile.get("title"),
            "image": profile.get("image_512")
        }
    except Exception as e:
        print(f"❌ Failed to get user info: {e}")
        return {"display_name": "there"}
    
# GPT response generator
def generate_gpt_table(user_text, file_path=None, user_comments=None):
    system_prompt = """You are a financial assistant at MMG (Multi Media Group). Your task is to extract invoice data
from vendor Statements of Account (SOAs) uploaded as PDF documents. Your output must follow these strict rules:

1. Output only a markdown table and nothing else — no explanations, summaries, or introductions.
2. The table must contain exactly four columns with these headers (case-sensitive):
   | Date | Invoice Number | Amount | Remaining Amount |
3. Each row must represent one invoice, with its date, invoice number, amount, and remaining amount.
4. Dates must be formatted as DD/Mon/YYYY (e.g., 28/Dec/2024).
5. Amounts must be raw numbers only — no commas, no currency symbols.
6. For extracting the correct Invoice Number column, apply the following logic:
   - Prioritize columns named (in order): "Invoice Number" → "Document No" → "External Document No"
   - Within those columns, extract values that follow this pattern: an uppercase letter prefix followed by digits or special characters, e.g.:
     • INVMMT-24-279
     • S-INV+-04868
     • PINV-004934
   - Ignore suffixes or extra labels like "May 2022" that come **after** the actual invoice number.
7. At the end of the SOA, there may be additional charges such as interest or finance fees that do not have a `Date` or `Invoice Number`. Only include such a row **if the row is clearly labeled with the word "interest" (case-insensitive)** in the original document. If detected, include a single additional row at the bottom of the table like this:
   | - | INTEREST | <amount> | <remaining amount> |
   Be VERY alert of this situation, as it can happen in any SOA. But do not confuse it for an actual invoice row or the total row at the bottom.
8. The markdown table must begin and end with pipe (`|`) characters for every row, including the header.
9. Do not include any TOTAL row. Exclude any totals or summaries at the bottom of the table.
10. If any row has missing or unclear data (and is not an interest charge), skip it — do not guess or add placeholders.
11. If there is no separate "Remaining Amount" column, assume the "Amount" is also the "Remaining Amount" and duplicate that value in both columns.
"""

    try:
        # Upload file to OpenAI
        if file_path:
            file_upload = client.files.create(
                file=open(file_path, 'rb'),
                purpose="user_data"
            )
            file_id = file_upload.id

            print(f"📤 Uploaded file to OpenAI: {file_id}")

            response = client.responses.create(
            model="gpt-4.1",
            input=[
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "input_file", "file_id": file_id},
                        {"type": "input_text", "text": user_text}
                    ] + (
                        [{"type": "input_text", "text": f"Additional context: {user_comments}"}]
                        if user_comments else []
                    )
                }
            ],
            store=False  # Do not store the file permanently
        )
            client.files.delete(file_id)

            return response.output_text
        else:
            return "No file provided."

    except Exception as e:
        print("❌ OpenAI file+response call failed:", e)
        return "Sorry, I couldn't process the document."

# Respond to @mentions in channels
@slack_app.event("app_mention")
def handle_mention(event, say):
    user_text = event['text']
    user = event['user']

    def respond():
        # context = retrieve_context(user_text)
        answer = generate_gpt_table(user_text, 0)
        say(text=f"<@{user}> {answer}")

    threading.Thread(target=respond).start()

@slack_app.event("message")
def handle_file_dm(event, say):
    if "rules" in event.get("text", "").lower():
        rules_message = """
📘 *Vendor Reconciliation Rules*

To use the vendor reconciliation bot, follow these steps:

1. 🧾 *Upload two files in a direct message*:
   • One *vendor* file (include "vendor" in the file name)
   • One *SOA* file (include "soa" in the file name)

2. ✍️ *Add a keyword in the message* to indicate parsing mode:
   • `clean` – for structured SOA Excel files  
   • `dirty` – for unstructured SOA Excel/PDF files (uses GPT)

3. ✅ The bot will:
   • Extract, compare, and reconcile invoices
   • Print matched, partially matched, and unmatched rows
   • Generate and save a filled Excel reconciliation sheet

💡 *Advanced - Only for `dirty` mode*:
   • If you add a message **on the next line** after `dirty`, it will be passed as *contextual comments to GPT*.
   • Example:
     ```
     dirty
     This SOA includes finance charges at the bottom and missing invoice dates.```

   This helps GPT extract better tables when the SOA format is messy.

—
If you're unsure, start with `dirty` mode — it's safer for unstructured files.
"""
        say(rules_message)
        return
    
    elif event.get("channel_type") == "im" and "files" in event:
        print("📥 File upload detected in DM")
        user = event["user"]
        channel = event["channel"]

        now = datetime.now()

        for file_info in event["files"]:
            download_slack_file(file_info, user)

        # Continue only when both files are present
        if not all(key in user_file_store[user] for key in ("vendor", "soa")):
            say("📄 Waiting for both SOA and Vendor files...")
            return

        vendor_path = user_file_store[user]["vendor"]
        soa_path = user_file_store[user]["soa"]

        message_lines = event.get("text", "").splitlines()
        message_flags = [line.lower().strip() for line in message_lines if line.strip()]
        is_clean = any("clean" in line for line in message_flags)
        is_dirty = any("dirty" in line for line in message_flags)

        user_comments = None
        if is_dirty and len(message_flags) > 1:
            # Assume anything after 'dirty' is a comment
            dirty_index = next(i for i, line in enumerate(message_flags) if "dirty" in line)
            user_comments = "\n".join(message_flags[dirty_index + 1:])

        if is_clean:
            process_clean_vendor_reconciliation(vendor_path, soa_path, say, channel)
        elif is_dirty:
            process_dirty_vendor_reconciliation(vendor_path, soa_path, say, channel, user_comments)
        else:
            say("❓ Please specify whether the files are `clean` or `dirty` in your message.")
            return
    


# Slack endpoint verification and event relay
@flask_app.route("/slack/events", methods=["POST"])
def slack_events():
    if request.json and "challenge" in request.json:
        return jsonify({"challenge": request.json["challenge"]})
    return handler.handle(request)

if __name__ == "__main__":
    init_db()
    flask_app.run(port=3000)
